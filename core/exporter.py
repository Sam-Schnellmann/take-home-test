import io
import json
import zipfile
 
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
 
from config import (
    PASS, REVIEW, FAIL,
    FIELD_LABELS,
    EXPORT_JSON_NAME, EXPORT_CSV_NAME,
    EXPORT_XLSX_NAME,
    XLSX_COLOR_FAIL, XLSX_COLOR_REVIEW,
)
 
 
# Flatten
def _flatten(result: dict) -> dict:
    # Convert a validation result into a flat row for Excel.
    row = {
        "Filename":    result.get("filename", ""),
        "Timestamp":   result.get("timestamp", ""),
        "Overall":     result.get("overall", ""),
        "Explanation": result.get("explanation", ""),
    }
 
    # Big 3 required fields
    for field in ["brand_name", "abv", "government_warning"]:
        r     = result.get(field, {})
        label = FIELD_LABELS.get(field, field)   # e.g. "Brand Name"
        row[f"{label} Status"]    = r.get("status", "")
        row[f"{label} Extracted"] = r.get("extracted", "")
        row[f"{label} Expected"]  = r.get("expected", "")
        row[f"{label} Note"]      = r.get("message", "")
 
    # Secondary fields
    for field, r in result.get("secondary", {}).items():
        label = FIELD_LABELS.get(field, field)
        row[f"{label} Status"]    = r.get("status", "")
        row[f"{label} Extracted"] = r.get("extracted", "")
        row[f"{label} Note"]      = r.get("message", "")
 
    return row
 
 
# JSON
def _build_json(results: list[dict]) -> bytes:
    return json.dumps(results, indent=2).encode("utf-8")
 
 
# CSV
def _build_csv(results: list[dict]) -> bytes:
    rows = [_flatten(r) for r in results]
    df   = pd.DataFrame(rows)
    buf  = io.StringIO()
    df.to_csv(buf, index=False)
    return buf.getvalue().encode("utf-8")
 
 
# XLSX
def _build_xlsx(results: list[dict]) -> bytes:
    rows = [_flatten(r) for r in results]
    if not rows:
        rows = [{}]
 
    wb = Workbook()
    ws = wb.active
    ws.title = "TTB Results"
 
    headers = list(rows[0].keys())
 
    # Header row styling
    header_fill      = PatternFill("solid", fgColor="FF2C3E50")  # dark navy
    header_font      = Font(bold=True, color="FFFFFFFF")         # white bold
 
    for col_idx, header in enumerate(headers, start=1):
        cell           = ws.cell(row=1, column=col_idx, value=header)
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
 
    # Data rows
    fill_fail   = PatternFill("solid", fgColor=XLSX_COLOR_FAIL)
    fill_review = PatternFill("solid", fgColor=XLSX_COLOR_REVIEW)
 
    for row_idx, row_data in enumerate(rows, start=2):
        overall = row_data.get("Overall", "")
        if overall == FAIL:
            row_fill = fill_fail
        elif overall == REVIEW:
            row_fill = fill_review
        else:
            row_fill = None
 
        for col_idx, header in enumerate(headers, start=1):
            cell           = ws.cell(row=row_idx, column=col_idx, value=row_data.get(header, ""))
            cell.alignment = Alignment(wrap_text=True)
            if row_fill:
                cell.fill = row_fill
 
    # Auto-size columns
    for col in ws.columns:
        max_len = max(
            (len(str(cell.value)) if cell.value else 0 for cell in col),
            default=0,
        )
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)
 
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
 
 
# ZIP folder for the three exports
def build_export_zip(results: list[dict]) -> bytes:
    """
    Build an in-memory ZIP containing JSON, CSV, and XLSX.
    Returns raw bytes ready for st.download_button.
    """
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr(EXPORT_JSON_NAME, _build_json(results))
        zf.writestr(EXPORT_CSV_NAME,  _build_csv(results))
        zf.writestr(EXPORT_XLSX_NAME, _build_xlsx(results))
    return buf.getvalue()